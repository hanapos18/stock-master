"""판매 엑셀 일괄 업로드 통합 테스트"""
import re
import requests
import pymysql
import os
from io import BytesIO
from openpyxl import Workbook
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))

BASE = "http://localhost:5556"
DB_CFG = dict(
    host=os.getenv("DB_HOST", "localhost"),
    port=int(os.getenv("DB_PORT", "3306")),
    user=os.getenv("DB_USER", "root"),
    password=os.getenv("DB_PASSWORD", ""),
    database=os.getenv("DB_NAME", "stock_master"),
    charset="utf8mb4",
    cursorclass=pymysql.cursors.DictCursor,
)


def db_query(sql, params=None):
    conn = pymysql.connect(**DB_CFG)
    cur = conn.cursor()
    cur.execute(sql, params)
    result = cur.fetchall()
    conn.close()
    return result


def db_one(sql, params=None):
    rows = db_query(sql, params)
    return rows[0] if rows else None


results = []


def check(name, condition):
    status = "PASS" if condition else "FAIL"
    results.append((name, status))
    print(f"  [{status}] {name}")


print("=== 판매 엑셀 일괄 업로드 테스트 시작 ===\n")

# 0. 테스트용 상품 코드 확인
biz = db_one("SELECT id FROM stk_businesses LIMIT 1")
biz_id = biz["id"]
products = db_query(
    "SELECT code, name, sell_price FROM stk_products WHERE business_id = %s LIMIT 3", (biz_id,)
)
if len(products) < 2:
    print("테스트용 상품이 2개 이상 필요합니다.")
    exit(1)
p1 = products[0]
p2 = products[1]
print(f"   상품1: {p1['code']} - {p1['name']} ({p1['sell_price']})")
print(f"   상품2: {p2['code']} - {p2['name']} ({p2['sell_price']})")

# 1. 로그인
print("\n1. 로그인")
s = requests.Session()
r = s.post(f"{BASE}/login", data={"username": "admin", "password": "admin123"}, allow_redirects=True)
check("로그인 성공", r.status_code == 200)

# 2. 판매 엑셀 템플릿 다운로드
print("\n2. 판매 엑셀 템플릿 다운로드")
r = s.get(f"{BASE}/sales/excel/template")
check("템플릿 다운로드 성공", r.status_code == 200)
check("xlsx 파일 형식", "spreadsheet" in r.headers.get("content-type", ""))

# 3. 업로드 페이지 접근
print("\n3. 업로드 페이지 접근")
r = s.get(f"{BASE}/sales/excel/upload")
check("업로드 페이지 접근", r.status_code == 200)
check("Download Template 링크 존재", "Download Template" in r.text)
check("Upload & Preview 버튼 존재", "Upload" in r.text)

# 4. 테스트 엑셀 파일 생성 & 업로드
print("\n4. 테스트 엑셀 파일 생성 & 업로드")
wb = Workbook()
ws = wb.active
ws.title = "Sales"
ws.append(["Sale Date*", "Customer Name", "Product Code*", "Product Name", "Quantity*", "Unit Price*", "Memo"])
ws.append(["2026-02-12", "Test Customer A", p1["code"], p1["name"], 2, float(p1["sell_price"]), "Excel test"])
ws.append(["2026-02-12", "Test Customer A", p2["code"], p2["name"], 3, float(p2["sell_price"]), "Excel test"])
ws.append(["2026-02-13", "Test Customer B", p1["code"], p1["name"], 1, float(p1["sell_price"]), "Another sale"])
buf = BytesIO()
wb.save(buf)
buf.seek(0)

# 판매 수 기록 (이전)
before_count = db_one("SELECT COUNT(*) as cnt FROM stk_sales WHERE business_id = %s", (biz_id,))["cnt"]

r = s.post(f"{BASE}/sales/excel/upload", files={"excel_file": ("test_sales.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
check("엑셀 업로드 성공", r.status_code == 200)
check("미리보기 페이지 표시", "Preview" in r.text or "preview" in r.text.lower())
check("2건 판매 그룹 표시", "2 Sale" in r.text or "Sale #1" in r.text)
check("상품 코드 표시", p1["code"] in r.text)
check("고객명 표시", "Test Customer A" in r.text)

# 5. Draft로 일괄 처리
print("\n5. Draft로 일괄 처리")
# 미리보기 HTML에서 grouped_data hidden field 추출
grouped_match = re.search(r'name="grouped_data"\s+value="([^"]*)"', r.text)
grouped_data = ""
if grouped_match:
    import html
    grouped_data = html.unescape(grouped_match.group(1))
check("grouped_data 추출 성공", len(grouped_data) > 10)

r = s.post(f"{BASE}/sales/excel/process", data={"auto_confirm": "0", "grouped_data": grouped_data}, allow_redirects=True)
check("일괄 처리 성공", r.status_code == 200)
check("성공 메시지 (created as draft)", "created as draft" in r.text or "success" in r.text.lower())

after_count = db_one("SELECT COUNT(*) as cnt FROM stk_sales WHERE business_id = %s", (biz_id,))["cnt"]
check(f"판매 2건 생성됨 ({before_count} -> {after_count})", after_count == before_count + 2)

# DB에서 생성된 판매 확인
new_sales = db_query(
    "SELECT * FROM stk_sales WHERE business_id = %s AND memo LIKE '%%Excel Upload%%' ORDER BY id DESC LIMIT 2",
    (biz_id,),
)
check("Excel Upload 메모 포함", len(new_sales) >= 2)
if new_sales:
    check("첫번째 판매 status = draft", new_sales[0]["status"] == "draft")

# 6. Confirm(FEFO) 일괄 처리 테스트
print("\n6. Confirm(FEFO) 일괄 처리 테스트")
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Sales"
ws2.append(["Sale Date*", "Customer Name", "Product Code*", "Product Name", "Quantity*", "Unit Price*", "Memo"])
ws2.append(["2026-02-13", "FEFO Test Client", p1["code"], p1["name"], 1, float(p1["sell_price"]), "FEFO test"])
buf2 = BytesIO()
wb2.save(buf2)
buf2.seek(0)

before_count2 = db_one("SELECT COUNT(*) as cnt FROM stk_sales WHERE business_id = %s", (biz_id,))["cnt"]

r = s.post(f"{BASE}/sales/excel/upload", files={"excel_file": ("test_fefo.xlsx", buf2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
check("FEFO 엑셀 업로드 성공", r.status_code == 200)

# grouped_data 추출
grouped_match2 = re.search(r'name="grouped_data"\s+value="([^"]*)"', r.text)
grouped_data2 = ""
if grouped_match2:
    import html
    grouped_data2 = html.unescape(grouped_match2.group(1))

r = s.post(f"{BASE}/sales/excel/process", data={"auto_confirm": "1", "grouped_data": grouped_data2}, allow_redirects=True)
check("FEFO 일괄 확인 처리 성공", r.status_code == 200)

after_count2 = db_one("SELECT COUNT(*) as cnt FROM stk_sales WHERE business_id = %s", (biz_id,))["cnt"]
check(f"판매 1건 추가 ({before_count2} -> {after_count2})", after_count2 == before_count2 + 1)

fefo_sale = db_one(
    "SELECT * FROM stk_sales WHERE business_id = %s AND memo LIKE '%%FEFO test%%' ORDER BY id DESC LIMIT 1",
    (biz_id,),
)
if fefo_sale:
    check("FEFO 판매 status = confirmed", fefo_sale["status"] == "confirmed")
else:
    check("FEFO 판매 찾기", False)

# 7. 잘못된 상품코드 에러 처리
print("\n7. 잘못된 상품코드 에러 처리")
wb3 = Workbook()
ws3 = wb3.active
ws3.title = "Sales"
ws3.append(["Sale Date*", "Customer Name", "Product Code*", "Product Name", "Quantity*", "Unit Price*", "Memo"])
ws3.append(["2026-02-13", "Error Test", "INVALID-CODE-999", "Nothing", 1, 1000, ""])
buf3 = BytesIO()
wb3.save(buf3)
buf3.seek(0)

r = s.post(f"{BASE}/sales/excel/upload", files={"excel_file": ("test_error.xlsx", buf3, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
check("에러 감지 (잘못된 코드)", "not found" in r.text.lower() or "error" in r.text.lower())

# 8. 판매 목록에서 Excel Upload 버튼 확인
print("\n8. 판매 목록 버튼 확인")
r = s.get(f"{BASE}/sales/")
check("판매 목록 접근", r.status_code == 200)
check("Excel Upload 버튼 존재", "Excel Upload" in r.text)

# 결과 요약
print("\n" + "=" * 50)
passed = sum(1 for _, s in results if s == "PASS")
failed = sum(1 for _, s in results if s == "FAIL")
print(f"결과: {passed} PASS / {failed} FAIL / 총 {len(results)} 건")
if failed > 0:
    print("\n실패 항목:")
    for name, status in results:
        if status == "FAIL":
            print(f"  - {name}")
print("=" * 50)
