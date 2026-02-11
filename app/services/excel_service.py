"""엑셀 내보내기/가져오기 서비스"""
from typing import List, Dict, Tuple
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def generate_excel_report(title: str, headers: List[str], rows: List[List],
                          column_widths: List[int] = None) -> BytesIO:
    """엑셀 리포트를 생성하고 BytesIO 스트림으로 반환합니다."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31]
    _write_title_row(sheet, title)
    _write_header_row(sheet, headers)
    _write_data_rows(sheet, rows, start_row=3)
    _set_column_widths(sheet, headers, column_widths)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _write_title_row(sheet, title: str) -> None:
    """제목 행을 작성합니다."""
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1)
    cell = sheet.cell(row=1, column=1, value=title)
    cell.font = Font(bold=True, size=14)


def _write_header_row(sheet, headers: List[str]) -> None:
    """헤더 행을 작성합니다."""
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _write_data_rows(sheet, rows: List[List], start_row: int = 3) -> None:
    """데이터 행들을 작성합니다."""
    for row_idx, row_data in enumerate(rows, start_row):
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"


def _set_column_widths(sheet, headers: List[str], widths: List[int] = None) -> None:
    """열 너비를 설정합니다."""
    for col_idx, header in enumerate(headers, 1):
        if widths and col_idx <= len(widths):
            width = widths[col_idx - 1]
        else:
            width = max(len(str(header)) + 4, 12)
        sheet.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = width


# ── 엑셀 가져오기 (Import) ──

PRODUCT_TEMPLATE_HEADERS = [
    "Code*", "Name*", "Barcode", "Category", "Supplier",
    "Unit", "Buy Price", "Sell Price", "Min Stock", "Max Stock",
    "Storage Location", "Description",
]

PRODUCT_TEMPLATE_WIDTHS = [14, 25, 18, 18, 18, 10, 14, 14, 12, 12, 20, 30]

PRODUCT_SAMPLE_ROWS = [
    ["P0001", "Rice 10kg", "8801234567890", "Food", "ABC Supplier",
     "bag", 25000, 30000, 5, 100, "Shelf A-1", "Premium rice"],
    ["P0002", "Soy Sauce 1L", "8809876543210", "Sauce", "",
     "bottle", 3500, 5000, 10, 50, "Shelf B-2", ""],
]


def generate_product_template() -> BytesIO:
    """상품 업로드용 엑셀 템플릿을 생성합니다."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    _write_template_instructions(workbook)
    _write_template_header_row(sheet, PRODUCT_TEMPLATE_HEADERS)
    _write_data_rows(sheet, PRODUCT_SAMPLE_ROWS, start_row=2)
    _set_template_column_widths(sheet, PRODUCT_TEMPLATE_HEADERS, PRODUCT_TEMPLATE_WIDTHS)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def parse_product_excel(file_stream: BytesIO) -> Tuple[List[Dict], List[str]]:
    """업로드된 엑셀 파일에서 상품 데이터를 파싱합니다.
    Returns: (parsed_rows, errors)
    """
    workbook = load_workbook(file_stream, read_only=True, data_only=True)
    sheet = _find_data_sheet(workbook)
    header_map = _build_header_map(sheet)
    if not header_map:
        return [], ["Invalid template: headers not found in first row"]
    rows: List[Dict] = []
    errors: List[str] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if _is_empty_row(row):
            continue
        parsed, row_errors = _parse_product_row(row, header_map, row_idx)
        if row_errors:
            errors.extend(row_errors)
        else:
            rows.append(parsed)
    workbook.close()
    return rows, errors


def _find_data_sheet(workbook):
    """데이터가 있는 시트를 찾습니다. 'Products' 시트 우선, 없으면 Instructions가 아닌 첫 시트."""
    if "Products" in workbook.sheetnames:
        return workbook["Products"]
    for name in workbook.sheetnames:
        if name.lower() != "instructions":
            return workbook[name]
    return workbook.active


def _build_header_map(sheet) -> Dict[str, int]:
    """첫 번째 행에서 헤더 매핑을 생성합니다."""
    header_map: Dict[str, int] = {}
    normalize_map = {
        "code": "code", "code*": "code",
        "name": "name", "name*": "name",
        "barcode": "barcode",
        "category": "category",
        "supplier": "supplier",
        "unit": "unit",
        "buy price": "unit_price", "buyprice": "unit_price", "unit price": "unit_price",
        "sell price": "sell_price", "sellprice": "sell_price",
        "min stock": "min_stock", "minstock": "min_stock",
        "max stock": "max_stock", "maxstock": "max_stock",
        "storage location": "storage_location", "location": "storage_location",
        "description": "description",
    }
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first_row:
        return {}
    for col_idx, cell_value in enumerate(first_row):
        if cell_value is None:
            continue
        key = str(cell_value).strip().lower().replace("_", " ")
        if key in normalize_map:
            header_map[normalize_map[key]] = col_idx
    if "code" not in header_map or "name" not in header_map:
        return {}
    return header_map


def _is_empty_row(row: tuple) -> bool:
    """빈 행인지 확인합니다."""
    return all(v is None or str(v).strip() == "" for v in row)


def _parse_product_row(row: tuple, header_map: Dict[str, int],
                       row_idx: int) -> Tuple[Dict, List[str]]:
    """한 행의 상품 데이터를 파싱합니다."""
    errors: List[str] = []
    def get_val(field: str, default=""):
        idx = header_map.get(field)
        if idx is None or idx >= len(row) or row[idx] is None:
            return default
        return str(row[idx]).strip()
    code = get_val("code")
    name = get_val("name")
    if not code:
        errors.append(f"Row {row_idx}: Code is required")
    if not name:
        errors.append(f"Row {row_idx}: Name is required")
    unit_price = _parse_number(get_val("unit_price", "0"), f"Row {row_idx} Buy Price", errors)
    sell_price = _parse_number(get_val("sell_price", "0"), f"Row {row_idx} Sell Price", errors)
    min_stock = _parse_number(get_val("min_stock", "0"), f"Row {row_idx} Min Stock", errors)
    max_stock_str = get_val("max_stock", "")
    max_stock = _parse_number(max_stock_str, f"Row {row_idx} Max Stock", errors) if max_stock_str else None
    return {
        "code": code,
        "name": name,
        "barcode": get_val("barcode"),
        "category_name": get_val("category"),
        "supplier_name": get_val("supplier"),
        "unit": get_val("unit", "ea") or "ea",
        "unit_price": unit_price,
        "sell_price": sell_price,
        "min_stock": min_stock,
        "max_stock": max_stock,
        "storage_location": get_val("storage_location"),
        "description": get_val("description"),
    }, errors


def _parse_number(value: str, label: str, errors: List[str]) -> float:
    """숫자 문자열을 float로 변환합니다."""
    if not value:
        return 0.0
    try:
        cleaned = value.replace(",", "")
        return float(cleaned)
    except ValueError:
        errors.append(f"{label}: '{value}' is not a valid number")
        return 0.0


def _write_template_instructions(workbook: Workbook) -> None:
    """템플릿에 안내 시트를 추가합니다."""
    sheet = workbook.create_sheet("Instructions")
    instructions = [
        ("Product Upload Template Instructions", ""),
        ("", ""),
        ("1. Fill in the 'Products' sheet with your product data.", ""),
        ("2. Code* and Name* are required fields.", ""),
        ("3. Category and Supplier names must match existing records.", ""),
        ("4. If a product with the same Code already exists, it will be updated.", ""),
        ("5. Delete the sample rows before uploading.", ""),
        ("6. Do not change the header row.", ""),
        ("", ""),
        ("Column Descriptions:", ""),
        ("Code*", "Unique product code (required)"),
        ("Name*", "Product name (required)"),
        ("Barcode", "Product barcode (optional)"),
        ("Category", "Category name - must exist in system"),
        ("Supplier", "Supplier name - must exist in system"),
        ("Unit", "Unit of measure (ea, kg, box, etc.) Default: ea"),
        ("Buy Price", "Purchase/buy price. Default: 0"),
        ("Sell Price", "Selling price. Default: 0"),
        ("Min Stock", "Minimum stock alert level. Default: 0"),
        ("Max Stock", "Maximum stock level (optional)"),
        ("Storage Location", "Physical storage location (optional)"),
        ("Description", "Product description (optional)"),
    ]
    for row_idx, (col_a, col_b) in enumerate(instructions, 1):
        sheet.cell(row=row_idx, column=1, value=col_a)
        sheet.cell(row=row_idx, column=2, value=col_b)
        if row_idx == 1:
            sheet.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
        if row_idx >= 11:
            sheet.cell(row=row_idx, column=1).font = Font(bold=True)
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 50
    workbook.move_sheet("Instructions", offset=-1)


def _write_template_header_row(sheet, headers: List[str]) -> None:
    """템플릿용 헤더를 row 1에 작성합니다."""
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _set_template_column_widths(sheet, headers: List[str], widths: List[int]) -> None:
    """템플릿 시트의 열 너비를 설정합니다."""
    for col_idx, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


# ── 매입(Purchase) 엑셀 가져오기 ──

PURCHASE_TEMPLATE_HEADERS = [
    "Purchase Date*", "Supplier", "Product Code*", "Product Name",
    "Quantity*", "Unit Price*", "Memo",
]
PURCHASE_TEMPLATE_WIDTHS = [16, 20, 14, 25, 12, 14, 25]
PURCHASE_SAMPLE_ROWS = [
    ["2026-02-11", "ABC Supplier", "P0001", "Rice 10kg", 10, 25000, "Monthly order"],
    ["2026-02-11", "ABC Supplier", "P0002", "Soy Sauce 1L", 20, 3500, "Monthly order"],
    ["2026-02-12", "XYZ Mart", "P0003", "Sugar 1kg", 5, 2000, ""],
]


def generate_purchase_template() -> BytesIO:
    """매입 업로드용 엑셀 템플릿을 생성합니다."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Purchases"
    _write_purchase_instructions(workbook)
    _write_template_header_row(sheet, PURCHASE_TEMPLATE_HEADERS)
    _write_data_rows(sheet, PURCHASE_SAMPLE_ROWS, start_row=2)
    _set_template_column_widths(sheet, PURCHASE_TEMPLATE_HEADERS, PURCHASE_TEMPLATE_WIDTHS)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def parse_purchase_excel(file_stream: BytesIO) -> Tuple[List[Dict], List[str]]:
    """매입 엑셀 파일을 파싱합니다. 같은 날짜+공급처+메모를 하나의 매입으로 그룹핑."""
    workbook = load_workbook(file_stream, read_only=True, data_only=True)
    sheet = _find_data_sheet(workbook)
    header_map = _build_purchase_header_map(sheet)
    if not header_map:
        return [], ["Invalid template: required headers not found"]
    rows: List[Dict] = []
    errors: List[str] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if _is_empty_row(row):
            continue
        parsed, row_errors = _parse_purchase_row(row, header_map, row_idx)
        if row_errors:
            errors.extend(row_errors)
        else:
            rows.append(parsed)
    workbook.close()
    return rows, errors


def _build_purchase_header_map(sheet) -> Dict[str, int]:
    """매입 헤더 매핑을 생성합니다."""
    normalize = {
        "purchase date": "purchase_date", "purchase date*": "purchase_date",
        "date": "purchase_date", "date*": "purchase_date",
        "supplier": "supplier", "supplier name": "supplier",
        "product code": "product_code", "product code*": "product_code",
        "code": "product_code", "code*": "product_code",
        "product name": "product_name", "name": "product_name",
        "quantity": "quantity", "quantity*": "quantity", "qty": "quantity",
        "unit price": "unit_price", "unit price*": "unit_price", "price": "unit_price",
        "memo": "memo", "note": "memo",
    }
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first_row:
        return {}
    header_map: Dict[str, int] = {}
    for col_idx, cell_value in enumerate(first_row):
        if cell_value is None:
            continue
        key = str(cell_value).strip().lower().replace("_", " ")
        if key in normalize:
            header_map[normalize[key]] = col_idx
    if "product_code" not in header_map or "quantity" not in header_map:
        return {}
    return header_map


def _parse_purchase_row(row: tuple, header_map: Dict[str, int],
                        row_idx: int) -> Tuple[Dict, List[str]]:
    """매입 한 행을 파싱합니다."""
    errors: List[str] = []
    def get_val(field: str, default=""):
        idx = header_map.get(field)
        if idx is None or idx >= len(row) or row[idx] is None:
            return default
        return str(row[idx]).strip()
    product_code = get_val("product_code")
    if not product_code:
        errors.append(f"Row {row_idx}: Product Code is required")
    quantity = _parse_number(get_val("quantity", "0"), f"Row {row_idx} Quantity", errors)
    if quantity <= 0 and not errors:
        errors.append(f"Row {row_idx}: Quantity must be greater than 0")
    unit_price = _parse_number(get_val("unit_price", "0"), f"Row {row_idx} Unit Price", errors)
    date_val = get_val("purchase_date")
    if date_val and "datetime" in str(type(row[header_map.get("purchase_date", 0)])).lower():
        date_val = row[header_map["purchase_date"]].strftime("%Y-%m-%d")
    return {
        "purchase_date": date_val,
        "supplier_name": get_val("supplier"),
        "product_code": product_code,
        "product_name": get_val("product_name"),
        "quantity": quantity,
        "unit_price": unit_price,
        "memo": get_val("memo"),
    }, errors


def _write_purchase_instructions(workbook: Workbook) -> None:
    """매입 템플릿 안내 시트를 추가합니다."""
    sheet = workbook.create_sheet("Instructions")
    instructions = [
        ("Purchase Upload Template Instructions", ""),
        ("", ""),
        ("1. Fill in the 'Purchases' sheet with your purchase data.", ""),
        ("2. Rows with the same Date + Supplier + Memo are grouped into one purchase.", ""),
        ("3. Product Code must match existing products in the system.", ""),
        ("4. Purchase Date format: YYYY-MM-DD", ""),
        ("5. Supplier name must match existing suppliers.", ""),
        ("6. Delete the sample rows before uploading.", ""),
    ]
    for row_idx, (col_a, col_b) in enumerate(instructions, 1):
        sheet.cell(row=row_idx, column=1, value=col_a)
        if row_idx == 1:
            sheet.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
    sheet.column_dimensions["A"].width = 60
    workbook.move_sheet("Instructions", offset=-1)


# ── 레시피(Recipe) 엑셀 가져오기 ──

RECIPE_TEMPLATE_HEADERS = [
    "Recipe Name*", "Product Code*", "Product Name",
    "Quantity*", "Unit", "Yield Qty", "Yield Unit", "Description",
]
RECIPE_TEMPLATE_WIDTHS = [25, 14, 25, 12, 10, 12, 12, 30]
RECIPE_SAMPLE_ROWS = [
    ["Fried Rice", "P0001", "Rice 10kg", 0.3, "kg", 1, "plate", "Basic fried rice"],
    ["Fried Rice", "P0002", "Soy Sauce 1L", 0.02, "L", 1, "plate", "Basic fried rice"],
    ["Fried Rice", "P0010", "Cooking Oil", 0.05, "L", 1, "plate", "Basic fried rice"],
    ["Miso Soup", "P0020", "Miso Paste", 0.03, "kg", 1, "bowl", ""],
    ["Miso Soup", "P0021", "Tofu", 0.1, "block", 1, "bowl", ""],
]


def generate_recipe_template() -> BytesIO:
    """레시피 업로드용 엑셀 템플릿을 생성합니다."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Recipes"
    _write_recipe_instructions(workbook)
    _write_template_header_row(sheet, RECIPE_TEMPLATE_HEADERS)
    _write_data_rows(sheet, RECIPE_SAMPLE_ROWS, start_row=2)
    _set_template_column_widths(sheet, RECIPE_TEMPLATE_HEADERS, RECIPE_TEMPLATE_WIDTHS)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def parse_recipe_excel(file_stream: BytesIO) -> Tuple[List[Dict], List[str]]:
    """레시피 엑셀 파일을 파싱합니다. 같은 Recipe Name을 하나의 레시피로 그룹핑."""
    workbook = load_workbook(file_stream, read_only=True, data_only=True)
    sheet = _find_data_sheet(workbook)
    header_map = _build_recipe_header_map(sheet)
    if not header_map:
        return [], ["Invalid template: required headers not found"]
    rows: List[Dict] = []
    errors: List[str] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if _is_empty_row(row):
            continue
        parsed, row_errors = _parse_recipe_row(row, header_map, row_idx)
        if row_errors:
            errors.extend(row_errors)
        else:
            rows.append(parsed)
    workbook.close()
    return rows, errors


def _build_recipe_header_map(sheet) -> Dict[str, int]:
    """레시피 헤더 매핑을 생성합니다."""
    normalize = {
        "recipe name": "recipe_name", "recipe name*": "recipe_name",
        "recipe": "recipe_name", "menu name": "recipe_name",
        "product code": "product_code", "product code*": "product_code",
        "code": "product_code", "code*": "product_code",
        "product name": "product_name", "name": "product_name",
        "quantity": "quantity", "quantity*": "quantity", "qty": "quantity",
        "unit": "unit",
        "yield qty": "yield_quantity", "yield quantity": "yield_quantity",
        "yield unit": "yield_unit",
        "description": "description",
    }
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first_row:
        return {}
    header_map: Dict[str, int] = {}
    for col_idx, cell_value in enumerate(first_row):
        if cell_value is None:
            continue
        key = str(cell_value).strip().lower().replace("_", " ")
        if key in normalize:
            header_map[normalize[key]] = col_idx
    if "recipe_name" not in header_map or "product_code" not in header_map:
        return {}
    return header_map


def _parse_recipe_row(row: tuple, header_map: Dict[str, int],
                      row_idx: int) -> Tuple[Dict, List[str]]:
    """레시피 한 행을 파싱합니다."""
    errors: List[str] = []
    def get_val(field: str, default=""):
        idx = header_map.get(field)
        if idx is None or idx >= len(row) or row[idx] is None:
            return default
        return str(row[idx]).strip()
    recipe_name = get_val("recipe_name")
    product_code = get_val("product_code")
    if not recipe_name:
        errors.append(f"Row {row_idx}: Recipe Name is required")
    if not product_code:
        errors.append(f"Row {row_idx}: Product Code is required")
    quantity = _parse_number(get_val("quantity", "0"), f"Row {row_idx} Quantity", errors)
    yield_qty = _parse_number(get_val("yield_quantity", "1"), f"Row {row_idx} Yield Qty", errors)
    return {
        "recipe_name": recipe_name,
        "product_code": product_code,
        "product_name": get_val("product_name"),
        "quantity": quantity,
        "unit": get_val("unit"),
        "yield_quantity": yield_qty if yield_qty > 0 else 1,
        "yield_unit": get_val("yield_unit", "ea") or "ea",
        "description": get_val("description"),
    }, errors


def _write_recipe_instructions(workbook: Workbook) -> None:
    """레시피 템플릿 안내 시트를 추가합니다."""
    sheet = workbook.create_sheet("Instructions")
    instructions = [
        ("Recipe Upload Template Instructions", ""),
        ("", ""),
        ("1. Fill in the 'Recipes' sheet with your recipe data.", ""),
        ("2. Rows with the same Recipe Name are grouped into one recipe.", ""),
        ("3. Product Code must match existing products in the system.", ""),
        ("4. Yield Qty/Unit: how much one batch makes (e.g. 1 plate).", ""),
        ("5. If recipe already exists (same name), it will be updated.", ""),
        ("6. Delete the sample rows before uploading.", ""),
    ]
    for row_idx, (col_a, col_b) in enumerate(instructions, 1):
        sheet.cell(row=row_idx, column=1, value=col_a)
        if row_idx == 1:
            sheet.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
    sheet.column_dimensions["A"].width = 60
    workbook.move_sheet("Instructions", offset=-1)
